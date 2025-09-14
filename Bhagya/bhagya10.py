# bom_map_app.py
import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook, Workbook
import difflib

st.set_page_config(layout="wide")
st.title("🔄 BOM Supplier Mapping (OLD → NEW)")

# ---------- helpers ----------
def detect_header_row_ws(ws, look_for=('mpn','manufacturer','coreel','part number')):
    """
    Find the header row index (1-based) by scanning first 30 rows for a candidate word.
    Returns header_row (int). Defaults to 1 if not found.
    """
    max_check = min(30, ws.max_row)
    for r in range(1, max_check+1):
        rowvals = []
        for c in range(1, min(300, ws.max_column)+1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                rowvals.append('')
            else:
                rowvals.append(str(v).strip().lower())
        # if any candidate substring occurs in the row, assume header row
        if any(any(x in cell for cell in rowvals) for x in look_for):
            return r
    return 1

def find_best_col_name(logical_name, available_cols):
    """
    Return the actual column name from available_cols that best matches logical_name (case-insensitive).
    If none close, return None.
    """
    low_map = {str(c).strip().lower(): c for c in available_cols}
    key = logical_name.strip().lower()
    if key in low_map:
        return low_map[key]
    # try close matches
    close = difflib.get_close_matches(key, list(low_map.keys()), n=1, cutoff=0.7)
    if close:
        return low_map[close[0]]
    return None

# Columns we want to transfer (logical names). We will match these to actual OLD BOM column names.
TRANSFER_COLS_LOGICAL = [
    "Supplier", "PO number", "Po qty", "Supplier part number",
    "Price", "Extended price", "Remarks", "ETA", "Currency",
    "Lead time", "Availability", "BCD", "unit price with BCD",
    "unit price in INR", "Extended price in INR"
]

# ---------- UI: upload ----------
col1, col2 = st.columns(2)
with col1:
    old_file = st.file_uploader("Upload OLD BOM (reference)", type=["xlsx"])
with col2:
    new_file = st.file_uploader("Upload NEW BOM (target)", type=["xlsx"])

if not old_file or not new_file:
    st.info("Upload both OLD and NEW BOM files (Excel) to proceed.")
    st.stop()

# read workbooks (openpyxl) from uploaded files
try:
    old_bytes = old_file.read()
    new_bytes = new_file.read()
    wb_old = load_workbook(filename=BytesIO(old_bytes), data_only=True)
    wb_new = load_workbook(filename=BytesIO(new_bytes), data_only=False)  # we want formula text from new
    ws_old = wb_old.active
    ws_new = wb_new.active
except Exception as e:
    st.error(f"Error loading workbooks: {e}")
    st.stop()

# detect header rows
hdr_old = detect_header_row_ws(ws_old)
hdr_new = detect_header_row_ws(ws_new)

st.info(f"Detected header row → OLD: {hdr_old}, NEW: {hdr_new}")

# Read pandas DataFrames using detected header rows (0-based header index = hdr-1)
old_df = pd.read_excel(BytesIO(old_bytes), header=hdr_old-1)
new_df = pd.read_excel(BytesIO(new_bytes), header=hdr_new-1)

# Normalize column lists (strip)
old_df.columns = [str(c).strip() for c in old_df.columns]
new_df.columns = [str(c).strip() for c in new_df.columns]

# find actual MPN column names in both files (case-insensitive / fuzzy)
mpn_candidates = ["mpn", "manufacturer part number", "part number", "mfr p/n", "coreel p/n"]
def detect_mpn_col_from_list(cols):
    lc = [str(c).strip().lower() for c in cols]
    for cand in mpn_candidates:
        if cand in lc:
            return cols[lc.index(cand)]
    # fuzzy fallback
    close = difflib.get_close_matches("mpn", lc, n=1, cutoff=0.6)
    if close:
        return cols[lc.index(close[0])]
    return None

old_mpn_col = detect_mpn_col_from_list(old_df.columns.tolist())
new_mpn_col = detect_mpn_col_from_list(new_df.columns.tolist())

if not old_mpn_col or not new_mpn_col:
    st.error("❌ Could not detect an MPN column in one or both files. Here are the headers:")
    st.write("OLD BOM headers:", old_df.columns.tolist())
    st.write("NEW BOM headers:", new_df.columns.tolist())
    st.stop()

st.success(f"MPN detected → OLD: '{old_mpn_col}'  NEW: '{new_mpn_col}'")

# find actual 'Alternate' column in OLD if exists (for alternate MPN)
alt_col = find_best_col_name("Alternate", old_df.columns)  # may return None

# Build mapping of transfer logical names -> actual old dataframe column names (if exists)
actual_old_col_for = {}
for logical in TRANSFER_COLS_LOGICAL:
    actual = find_best_col_name(logical, old_df.columns)
    # If not found, create a placeholder column in old_df (filled with None) so code downstream can handle
    if actual is None:
        old_df[logical] = None
        actual_old_col_for[logical] = logical
    else:
        actual_old_col_for[logical] = actual

# Build mapping dictionary from OLD BOM: key -> dict-of-transfer-values
# keys: MPN and Alternate (if present)
mapping = {}
for _, r in old_df.iterrows():
    key = r.get(old_mpn_col, None)
    if pd.isna(key):
        continue
    key = str(key).strip()
    # build values dict using actual_old_col_for mapping
    vals = {}
    for logical in TRANSFER_COLS_LOGICAL:
        actual_col = actual_old_col_for[logical]
        vals[logical] = r.get(actual_col, None) if actual_col in old_df.columns else None
    # set mapping for mpn if not present already
    if key not in mapping:
        mapping[key] = vals
    # also map alternate if available
    if alt_col:
        alt_val = r.get(alt_col, None)
        if pd.notna(alt_val):
            alt_val = str(alt_val).strip()
            if alt_val and alt_val not in mapping:
                mapping[alt_val] = vals

st.write(f"Loaded {len(mapping)} reference MPN rows from OLD BOM")

# Build header_map for NEW worksheet using openpyxl header row (to find column indexes)
header_map_new = {}
for c in range(1, ws_new.max_column + 1):
    val = ws_new.cell(row=hdr_new, column=c).value
    if val is not None:
        header_map_new[str(val).strip()] = c

# Prepare final header order:
new_cols = list(new_df.columns)
# determine exact strings for MPN and Manufacturer in new_df columns (we found new_mpn_col)
mpn_name = new_mpn_col
# detect manufacturer name in new_df (fuzzy)
manuf_name = find_best_col_name("Manufacturer", new_df.columns)
if manuf_name is None:
    st.error("❌ 'Manufacturer' column not found in NEW BOM headers. Please ensure it exists.")
    st.stop()

# Build final_headers: everything up to and including MPN, then TRANSFER_COLS_LOGICAL, then Manufacturer onward (excluding duplicates)
mpn_idx = new_cols.index(mpn_name)
manuf_idx = new_cols.index(manuf_name)
before = new_cols[:mpn_idx+1]  # includes MPN
after = new_cols[manuf_idx:]   # includes Manufacturer and everything after
# ensure we don't duplicate transfer cols if they already exist in new
after_filtered = [c for c in after if c not in TRANSFER_COLS_LOGICAL]
final_headers = before + TRANSFER_COLS_LOGICAL + after_filtered

# Build output workbook and worksheet
wb_out = Workbook()
ws_out = wb_out.active

# Write header row into output at same header row index
for col_i, header in enumerate(final_headers, start=1):
    ws_out.cell(row=hdr_new, column=col_i, value=header)

# Build reverse map: header name -> column index in input NEW sheet (if available)
# header_map_new already holds this; but keys might differ in case/spacing, so build normalized lookup
norm_header_map_new = {str(k).strip(): v for k, v in header_map_new.items()}

# Iterate data rows and populate values in ws_out
max_row_new = ws_new.max_row
for r in range(hdr_new + 1, max_row_new + 1):
    # read original row values for new sheet into a dict (header->value), use header_map_new
    orig_values = {}
    for hdr_name, col_idx in norm_header_map_new.items():
        orig_values[hdr_name] = ws_new.cell(row=r, column=col_idx).value

    # get mpn key (use the header string exactly as in new_df columns)
    mpn_val = None
    # find the column index for mpn_name in ws_new header map; if not found fallback to scanning orig_values keys for close match
    mpn_col_idx = norm_header_map_new.get(mpn_name)
    if mpn_col_idx:
        mpn_val = ws_new.cell(row=r, column=mpn_col_idx).value
    else:
        # fallback: search in orig_values by fuzzy match
        for k in orig_values.keys():
            if k.strip().lower() == mpn_name.strip().lower():
                mpn_val = orig_values[k]
                break

    mpn_key = str(mpn_val).strip() if mpn_val is not None else ""

    mapped_vals = mapping.get(mpn_key)

    # for each header in final_headers, write either mapped value (if transfer col) or original cell (copy formula/value)
    for c_idx, header in enumerate(final_headers, start=1):
        if header in TRANSFER_COLS_LOGICAL:
            # get mapped value if present
            if mapped_vals:
                v = mapped_vals.get(header, None)
            else:
                v = None
            # special handling for Remarks: prefer mapped remark if present; if not mapped and original remark empty -> "New Part"
            if header == "Remarks":
                # original remark from new file (if any)
                orig_rem = None
                # try to find original 'Remarks' in orig_values (case sensitive exact)
                if 'Remarks' in orig_values and orig_values['Remarks'] is not None:
                    orig_rem = orig_values['Remarks']
                elif 'Remarks' in new_df.columns:
                    # fallback to pandas new_df row value (index r - header_row - 1)
                    try:
                        idx = r - (hdr_new + 1)
                        if idx >= 0 and idx < len(new_df):
                            orig_rem = new_df.iloc[idx][ 'Remarks' ] if 'Remarks' in new_df.columns else None
                    except Exception:
                        orig_rem = None
                if mapped_vals and mapped_vals.get('Remarks') not in (None, ''):
                    v = mapped_vals.get('Remarks')
                else:
                    if (not mapped_vals) and (orig_rem is None or str(orig_rem).strip() == ""):
                        v = "New Part"
                    elif orig_rem is not None and (not mapped_vals):
                        v = orig_rem
            ws_out.cell(row=r, column=c_idx, value=v)
        else:
            # copy original cell (value or formula). We need to find original column index for this header.
            orig_col_idx = norm_header_map_new.get(header)
            if orig_col_idx:
                orig_cell = ws_new.cell(row=r, column=orig_col_idx)
                # Copy the cell value/formula
                ws_out.cell(row=r, column=c_idx, value=orig_cell.value)
            else:
                # header not in original new sheet (rare) => leave blank
                ws_out.cell(row=r, column=c_idx, value=None)

# Save to buffer and provide download
buffer = BytesIO()
wb_out.save(buffer)
buffer.seek(0)

st.success("✅ Mapping finished — download below")
st.download_button(
    "📥 Download Mapped NEW BOM (ERPU2_MAPPED.xlsx)",
    data=buffer.getvalue(),
    file_name="ERPU2_MAPPED.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

# Show small preview (pandas) for convenience
try:
    df_preview = pd.read_excel(BytesIO(buffer.getvalue()), header=hdr_new-1)
    st.subheader("Preview (first 40 rows)")
    st.dataframe(df_preview.head(40))
except Exception:
    st.info("Preview unavailable")
