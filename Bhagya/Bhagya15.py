# bom_map_app.py
import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import difflib

st.set_page_config(layout="wide")
st.title("🔄 BOM Supplier Mapping (OLD → NEW)")

# ---------- helpers ----------
def detect_header_row_ws(ws, look_for=('mpn','manufacturer','coreel','part number')):
    max_check = min(30, ws.max_row)
    for r in range(1, max_check+1):
        rowvals = []
        for c in range(1, min(300, ws.max_column)+1):
            v = ws.cell(row=r, column=c).value
            rowvals.append(str(v).strip().lower() if v else "")
        if any(any(x in cell for cell in rowvals) for x in look_for):
            return r
    return 1

def find_best_col_name(logical_name, available_cols):
    low_map = {str(c).strip().lower(): c for c in available_cols}
    key = logical_name.strip().lower()
    if key in low_map:
        return low_map[key]
    close = difflib.get_close_matches(key, list(low_map.keys()), n=1, cutoff=0.7)
    if close:
        return low_map[close[0]]
    return None

TRANSFER_COLS_LOGICAL = [
    "Supplier", "PO number", "Po qty", "Supplier part number",
    "Price", "Extended price", "Remarks", "ETA", "Currency",
    "Lead time", "Availability", "BCD", "unit price with BCD",
    "unit price in INR", "Extended price in INR"
]

# ---------- UI: upload ----------
col1, col2 = st.columns(2)
with col1:
    old_file = st.file_uploader("Upload OLD BOM (reference)", type=["xlsx"])
with col2:
    new_file = st.file_uploader("Upload NEW BOM (target)", type=["xlsx"])

if not old_file or not new_file:
    st.info("Upload both OLD and NEW BOM files (Excel) to proceed.")
    st.stop()

# ---------- load workbooks ----------
try:
    old_bytes = old_file.read()
    new_bytes = new_file.read()
    wb_old = load_workbook(filename=BytesIO(old_bytes), data_only=True)
    wb_new = load_workbook(filename=BytesIO(new_bytes), data_only=False)  # keep formulas
    ws_old = wb_old.active
    ws_new = wb_new.active
except Exception as e:
    st.error(f"Error loading workbooks: {e}")
    st.stop()

# detect header rows
hdr_old = detect_header_row_ws(ws_old)
hdr_new = detect_header_row_ws(ws_new)
st.info(f"Detected header row → OLD: {hdr_old}, NEW: {hdr_new}")

# read into pandas for easy row lookups (pandas read happens BEFORE we insert columns)
old_df = pd.read_excel(BytesIO(old_bytes), header=hdr_old-1)
new_df = pd.read_excel(BytesIO(new_bytes), header=hdr_new-1)
old_df.columns = [str(c).strip() for c in old_df.columns]
new_df.columns = [str(c).strip() for c in new_df.columns]

# detect MPN columns (OLD & NEW)
mpn_candidates = ["mpn", "manufacturer part number", "part number", "mfr p/n", "coreel p/n"]
def detect_mpn_col_from_list(cols):
    lc = [str(c).strip().lower() for c in cols]
    for cand in mpn_candidates:
        if cand in lc:
            return cols[lc.index(cand)]
    close = difflib.get_close_matches("mpn", lc, n=1, cutoff=0.6)
    if close:
        return cols[lc.index(close[0])]
    return None

old_mpn_col = detect_mpn_col_from_list(old_df.columns.tolist())
new_mpn_col = detect_mpn_col_from_list(new_df.columns.tolist())
if not old_mpn_col or not new_mpn_col:
    st.error("❌ Could not detect an MPN column in one or both files.")
    st.write("OLD BOM headers:", old_df.columns.tolist())
    st.write("NEW BOM headers:", new_df.columns.tolist())
    st.stop()

st.success(f"MPN detected → OLD: '{old_mpn_col}'  NEW: '{new_mpn_col}'")

# detect Alternate column name in OLD (optional)
alt_col = find_best_col_name("Alternate", old_df.columns)

# ---------- build mapping from OLD (normalize keys to lowercase for robust matching) ----------
actual_old_col_for = {}
for logical in TRANSFER_COLS_LOGICAL:
    actual = find_best_col_name(logical, old_df.columns)
    if actual is None:
        old_df[logical] = None
        actual_old_col_for[logical] = logical
    else:
        actual_old_col_for[logical] = actual

mapping = {}
for _, r in old_df.iterrows():
    key = r.get(old_mpn_col, None)
    if pd.isna(key):
        continue
    key_str = str(key).strip().lower()
    vals = {}
    for logical in TRANSFER_COLS_LOGICAL:
        actual_col = actual_old_col_for[logical]
        vals[logical] = r.get(actual_col, None) if actual_col in old_df.columns else None
    # store mapping under lowercase key
    if key_str not in mapping:
        mapping[key_str] = vals
    # also map alternate if present
    if alt_col:
        alt_val = r.get(alt_col, None)
        if pd.notna(alt_val):
            alt_key = str(alt_val).strip().lower()
            if alt_key and alt_key not in mapping:
                mapping[alt_key] = vals

st.write(f"Loaded {len(mapping)} reference MPN rows from OLD BOM")

# ---------- INSERT MAPPED COLUMNS IMMEDIATELY AFTER COLUMN C (i.e., at index 4 / 'D') ----------
insert_at = 4  # after C
num_new_cols = len(TRANSFER_COLS_LOGICAL)
# Insert all mapping columns in one call so existing cells (formulas/styles) are shifted correctly
ws_new.insert_cols(insert_at, amount=num_new_cols)

# Highlight header fill for inserted headers
header_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

# Write headers for inserted columns
for j, h in enumerate(TRANSFER_COLS_LOGICAL):
    col_idx = insert_at + j
    ws_new.cell(row=hdr_new, column=col_idx, value=h)
    ws_new.cell(row=hdr_new, column=col_idx).fill = header_fill

# ---------- Fill inserted columns row-by-row using pandas new_df (which matched rows before insert) ----------
max_row_new = ws_new.max_row
n_rows_df = len(new_df)

# We'll use pandas new_df to read original per-row MPN and original 'Remarks' (if present).
for r in range(hdr_new + 1, max_row_new + 1):
    df_idx = r - (hdr_new + 1)  # zero-based index into new_df
    orig_row = new_df.iloc[df_idx] if 0 <= df_idx < n_rows_df else None

    # get MPN from original dataframe (robust)
    mpn_val = orig_row[new_mpn_col] if orig_row is not None else None
    mpn_key = str(mpn_val).strip().lower() if mpn_val is not None else ""

    mapped_vals = mapping.get(mpn_key)  # mapping keys are lowercased

    # original remark (if present in NEW BOM)
    orig_rem = None
    if orig_row is not None and 'Remarks' in new_df.columns:
        orig_rem = orig_row.get('Remarks', None)

    for j, logical in enumerate(TRANSFER_COLS_LOGICAL):
        c_idx = insert_at + j
        v = None
        if mapped_vals:
            # prefer mapped value if present (not NaN)
            mv = mapped_vals.get(logical)
            if pd.notna(mv):
                v = mv
            else:
                # no mapped value for this logical col
                if logical.lower() == "remarks":
                    # if mapped exists but no remark, fallback to original remark if present
                    if orig_rem is not None and str(orig_rem).strip() != "":
                        v = orig_rem
                    else:
                        # if mapping exists overall but remark missing, leave blank (or set "New Part"?)
                        # conservative: leave blank (user can override). If you want "New Part" when mapping exists but remark missing, change here.
                        v = None
                else:
                    v = None
        else:
            # no mapping for this MPN
            if logical.lower() == "remarks":
                # if original remark present, keep it; else mark New Part
                if orig_rem is not None and str(orig_rem).strip() != "":
                    v = orig_rem
                else:
                    v = "New Part"
            else:
                v = None

        # write the value into the inserted column cell
        ws_new.cell(row=r, column=c_idx, value=v)

# ---------- Save modified NEW workbook to buffer and provide download ----------
buffer = BytesIO()
wb_new.save(buffer)
buffer.seek(0)

st.success("✅ Mapping finished — inserted supplier columns after C and preserved all original formulas/formatting.")
st.download_button(
    "📥 Download Mapped NEW BOM (ERPU2_MAPPED.xlsx)",
    data=buffer.getvalue(),
    file_name="ERPU2_MAPPED.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
