# bom_map_app.py
import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
import difflib

st.set_page_config(layout="wide")
st.title("🔄 BOM Supplier Mapping (OLD → NEW)")

# ---------- helpers ----------
def detect_header_row_ws(ws, look_for=('mpn','manufacturer','coreel','part number')):
    max_check = min(30, ws.max_row)
    for r in range(1, max_check+1):
        rowvals = []
        for c in range(1, min(300, ws.max_column)+1):
            v = ws.cell(row=r, column=c).value
            rowvals.append(str(v).strip().lower() if v else "")
        if any(any(x in cell for cell in rowvals) for x in look_for):
            return r
    return 1

def find_best_col_name(logical_name, available_cols):
    low_map = {str(c).strip().lower(): c for c in available_cols}
    key = logical_name.strip().lower()
    if key in low_map:
        return low_map[key]
    close = difflib.get_close_matches(key, list(low_map.keys()), n=1, cutoff=0.7)
    if close:
        return low_map[close[0]]
    return None

TRANSFER_COLS_LOGICAL = [
    "Supplier", "PO number", "Po qty", "Supplier part number",
    "Price", "Extended price", "Remarks", "ETA", "Currency",
    "Lead time", "Availability", "BCD", "unit price with BCD",
    "unit price in INR", "Extended price in INR"
]

# ---------- UI ----------
col1, col2 = st.columns(2)
with col1:
    old_file = st.file_uploader("Upload OLD BOM (reference)", type=["xlsx"])
with col2:
    new_file = st.file_uploader("Upload NEW BOM (target)", type=["xlsx"])

if not old_file or not new_file:
    st.info("Upload both OLD and NEW BOM files (Excel) to proceed.")
    st.stop()

try:
    old_bytes = old_file.read()
    new_bytes = new_file.read()
    wb_old = load_workbook(filename=BytesIO(old_bytes), data_only=True)
    wb_new = load_workbook(filename=BytesIO(new_bytes), data_only=False)  # keep formulas
    ws_old = wb_old.active
    ws_new = wb_new.active
except Exception as e:
    st.error(f"Error loading workbooks: {e}")
    st.stop()

hdr_old = detect_header_row_ws(ws_old)
hdr_new = detect_header_row_ws(ws_new)
st.info(f"Detected header row → OLD: {hdr_old}, NEW: {hdr_new}")

old_df = pd.read_excel(BytesIO(old_bytes), header=hdr_old-1)
new_df = pd.read_excel(BytesIO(new_bytes), header=hdr_new-1)
old_df.columns = [str(c).strip() for c in old_df.columns]
new_df.columns = [str(c).strip() for c in new_df.columns]

# detect MPN column
mpn_candidates = ["mpn", "manufacturer part number", "part number", "mfr p/n", "coreel p/n"]
def detect_mpn_col_from_list(cols):
    lc = [str(c).strip().lower() for c in cols]
    for cand in mpn_candidates:
        if cand in lc:
            return cols[lc.index(cand)]
    close = difflib.get_close_matches("mpn", lc, n=1, cutoff=0.6)
    if close:
        return cols[lc.index(close[0])]
    return None

old_mpn_col = detect_mpn_col_from_list(old_df.columns.tolist())
new_mpn_col = detect_mpn_col_from_list(new_df.columns.tolist())
if not old_mpn_col or not new_mpn_col:
    st.error("❌ Could not detect an MPN column in one or both files.")
    st.write("OLD BOM headers:", old_df.columns.tolist())
    st.write("NEW BOM headers:", new_df.columns.tolist())
    st.stop()

st.success(f"MPN detected → OLD: '{old_mpn_col}'  NEW: '{new_mpn_col}'")

# detect Alternate col
alt_col = find_best_col_name("Alternate", old_df.columns)

# Build OLD mapping
actual_old_col_for = {}
for logical in TRANSFER_COLS_LOGICAL:
    actual = find_best_col_name(logical, old_df.columns)
    if actual is None:
        old_df[logical] = None
        actual_old_col_for[logical] = logical
    else:
        actual_old_col_for[logical] = actual

mapping = {}
for _, r in old_df.iterrows():
    key = r.get(old_mpn_col, None)
    if pd.isna(key):
        continue
    key = str(key).strip()
    vals = {}
    for logical in TRANSFER_COLS_LOGICAL:
        actual_col = actual_old_col_for[logical]
        vals[logical] = r.get(actual_col, None) if actual_col in old_df.columns else None
    if key not in mapping:
        mapping[key] = vals
    if alt_col:
        alt_val = r.get(alt_col, None)
        if pd.notna(alt_val):
            alt_val = str(alt_val).strip()
            if alt_val and alt_val not in mapping:
                mapping[alt_val] = vals

st.write(f"Loaded {len(mapping)} reference MPN rows from OLD BOM")

# ---------- Insert new supplier columns after CC (so starting CD) ----------
cc_index = None
for idx, cell in enumerate(ws_new[hdr_new], start=1):
    if str(cell.value).strip().lower() == "cc":
        cc_index = idx
        break

if not cc_index:
    st.error("❌ Could not find 'CC' column in NEW BOM header row.")
    st.stop()

target_start_col = cc_index + 1  # after CC → CD

for i, col in enumerate(TRANSFER_COLS_LOGICAL):
    insert_at = target_start_col + i
    ws_new.insert_cols(insert_at)
    ws_new.cell(row=hdr_new, column=insert_at, value=col)

# Fill data rows
max_row_new = ws_new.max_row
new_headers = [c.value for c in ws_new[hdr_new]]

try:
    mpn_col_idx = new_headers.index(new_mpn_col) + 1
except ValueError:
    st.error("❌ Could not find MPN column in NEW BOM after insert.")
    st.stop()

for r in range(hdr_new + 1, max_row_new + 1):
    mpn_val = ws_new.cell(row=r, column=mpn_col_idx).value
    mpn_key = str(mpn_val).strip() if mpn_val is not None else ""
    mapped_vals = mapping.get(mpn_key)

    for i, col in enumerate(TRANSFER_COLS_LOGICAL):
        c_idx = target_start_col + i
        v = mapped_vals.get(col) if mapped_vals else None
        if col.lower() == "remarks":
            orig_rem = ws_new.cell(row=r, column=c_idx).value
            if mapped_vals and mapped_vals.get("Remarks"):
                v = mapped_vals.get("Remarks")
            elif not mapped_vals and (orig_rem is None or str(orig_rem).strip() == ""):
                v = "New Part"
            elif not mapped_vals and orig_rem:
                v = orig_rem
        ws_new.cell(row=r, column=c_idx, value=v)

# ---------- Save result ----------
buffer = BytesIO()
wb_new.save(buffer)
buffer.seek(0)

st.success("✅ Mapping finished — download below")
st.download_button(
    "📥 Download Mapped NEW BOM (ERPU2_MAPPED.xlsx)",
    data=buffer.getvalue(),
    file_name="ERPU2_MAPPED.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
