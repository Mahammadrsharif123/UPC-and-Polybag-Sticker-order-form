import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO

st.title("📊 BOM Mapper (MPN → Supplier/PO Details Transfer)")

# Upload files
new_file = st.file_uploader("Upload NEW Excel (Target - ERPU2_NEW.xlsx)", type=["xlsx"])
old_file = st.file_uploader("Upload OLD Excel (Reference - ERPU2_OLD.xlsx)", type=["xlsx"])

if new_file and old_file:
    # Read both files
    new_df = pd.read_excel(new_file)
    old_df = pd.read_excel(old_file)

    st.success("✅ Both files uploaded successfully!")

    # Normalize column names (remove spaces, lowercase)
    new_df.columns = new_df.columns.str.strip()
    old_df.columns = old_df.columns.str.strip()

    # Columns to transfer
    transfer_cols = [
        "Supplier", "PO number", "PO qty", "Supplier part number", "Price",
        "Extended price", "Remarks", "ETA", "Currency", "Lead time", "Availability",
        "BCD", "unit price with BCD", "unit price in INR", "Extended price in INR"
    ]

    # Check available columns in OLD file
    missing = [c for c in transfer_cols if c not in old_df.columns]
    if missing:
        st.error(f"❌ Missing columns in OLD file: {missing}")
        st.stop()

    # Build mapping dict (MPN → row values)
    mapping = {}
    for _, row in old_df.iterrows():
        mpn = str(row.get("MPN", "")).strip()
        alt = str(row.get("Alternate MPN", "")).strip()
        if mpn:
            mapping[mpn] = row[transfer_cols].to_dict()
        if alt:
            mapping[alt] = row[transfer_cols].to_dict()

    # Create merged DataFrame
    merged_rows = []
    for _, row in new_df.iterrows():
        mpn = str(row.get("MPN", "")).strip()
        transfer_data = mapping.get(mpn, {c: None for c in transfer_cols})
        merged_row = {}

        for col in new_df.columns:
            merged_row[col] = row[col]

        # Insert between MPN and Manufacturer
        new_order = []
        for col in merged_row.keys():
            new_order.append(col)
            if col == "MPN":
                for tcol in transfer_cols:
                    merged_row[tcol] = transfer_data.get(tcol)
                new_order.extend(transfer_cols)
        merged_rows.append(merged_row)

    merged_df = pd.DataFrame(merged_rows)

    # Reorder columns correctly (MPN → TransferCols → Manufacturer)
    cols = list(new_df.columns)
    if "MPN" in cols and "Manufacturer" in cols:
        idx = cols.index("MPN")
        new_cols = cols[:idx+1] + transfer_cols + cols[idx+1:]
        merged_df = merged_df[new_cols]

    st.subheader("🔎 Preview of Mapped Data")
    st.dataframe(merged_df.head(50))

    # Save to Excel with formulas intact
    new_file.seek(0)
    wb = load_workbook(new_file)
    ws = wb.active

    # Find header row
    headers = [cell.value for cell in ws[1]]
    mpn_idx = headers.index("MPN") + 1
    manuf_idx = headers.index("Manufacturer") + 1

    # Insert new columns in Excel
    for i, col_name in enumerate(transfer_cols):
        ws.insert_cols(manuf_idx + i)
        ws.cell(row=1, column=manuf_idx + i, value=col_name)

    # Fill values row by row
    for r_idx, row in enumerate(merged_df.itertuples(index=False), start=2):
        for c_idx, col_name in enumerate(transfer_cols, start=mpn_idx + 1):
            ws.cell(row=r_idx, column=c_idx, value=getattr(row, col_name))

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    st.download_button(
        label="📥 Download Final Mapped File",
        data=buffer,
        file_name="ERPU2_MAPPED.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
