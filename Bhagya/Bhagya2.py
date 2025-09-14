import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO

st.title("📊 BOM Mapper (MPN → Supplier/PO Details Transfer)")

# Upload files
new_file = st.file_uploader("Upload NEW Excel (Target - ERPU2_NEW.xlsx)", type=["xlsx"])
old_file = st.file_uploader("Upload OLD Excel (Reference - ERPU2_OLD.xlsx)", type=["xlsx"])

if new_file and old_file:
    new_df = pd.read_excel(new_file)
    old_df = pd.read_excel(old_file)

    st.success("✅ Both files uploaded successfully!")

    # Normalize column names
    new_df.columns = new_df.columns.str.strip()
    old_df.columns = old_df.columns.str.strip()

    # Define required logical columns
    required_cols = {
        "Supplier": ["Supplier", "Vendor", "Supplier Name"],
        "PO number": ["PO number", "PO Number", "PO#"],
        "PO qty": ["PO qty", "PO Quantity", "Qty"],
        "Supplier part number": ["Supplier part number", "Supplier Part No", "Supplier Part#"],
        "Price": ["Price", "Unit Price"],
        "Extended price": ["Extended price", "Total Price"],
        "Remarks": ["Remarks", "Comments", "Notes"],
        "ETA": ["ETA", "Expected Date"],
        "Currency": ["Currency", "Curr"],
        "Lead time": ["Lead time", "LT"],
        "Availability": ["Availability", "Stock"],
        "BCD": ["BCD"],
        "unit price with BCD": ["unit price with BCD", "Price+BCD"],
        "unit price in INR": ["unit price in INR", "Price INR"],
        "Extended price in INR": ["Extended price in INR", "Total INR"]
    }

    # Map actual old_df columns
    col_map = {}
    missing_cols = []
    for logical, aliases in required_cols.items():
        found = None
        for alias in aliases:
            if alias in old_df.columns:
                found = alias
                break
        if found:
            col_map[logical] = found
        else:
            missing_cols.append(logical)

    if missing_cols:
        st.warning(f"⚠️ Could not find these columns in OLD file: {missing_cols}")
        st.write("Available columns in OLD file are:", list(old_df.columns))

    transfer_cols = list(col_map.keys())  # only mapped columns

    # Build mapping (MPN → transfer cols)
    mapping = {}
    for _, row in old_df.iterrows():
        mpn = str(row.get("MPN", "")).strip()
        alt = str(row.get("Alternate MPN", "")).strip()
        row_data = {c: row[col_map[c]] if c in col_map else None for c in transfer_cols}
        if mpn:
            mapping[mpn] = row_data
        if alt:
            mapping[alt] = row_data

    # Merge with new_df
    merged_rows = []
    for _, row in new_df.iterrows():
        mpn = str(row.get("MPN", "")).strip()
        transfer_data = mapping.get(mpn)

        if transfer_data:  # ✅ Found in old file
            merged_row = dict(row)
            for c in transfer_cols:
                merged_row[c] = transfer_data.get(c)
        else:  # ❌ Not found → mark as "New Part"
            merged_row = dict(row)
            for c in transfer_cols:
                merged_row[c] = None
            merged_row["Remarks"] = "New Part"

        merged_rows.append(merged_row)

    merged_df = pd.DataFrame(merged_rows)

    # Reorder: insert transfer cols between MPN and Manufacturer
    cols = list(new_df.columns)
    if "MPN" in cols and "Manufacturer" in cols:
        idx = cols.index("MPN")
        new_cols = cols[:idx+1] + transfer_cols + cols[idx+1:]
        merged_df = merged_df[new_cols]

    st.subheader("🔎 Preview of Mapped Data")
    st.dataframe(merged_df.head(50))

    # Write back into Excel
    new_file.seek(0)
    wb = load_workbook(new_file)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    mpn_idx = headers.index("MPN") + 1
    manuf_idx = headers.index("Manufacturer") + 1

    # Insert columns in Excel
    for i, col_name in enumerate(transfer_cols):
        ws.insert_cols(manuf_idx + i)
        ws.cell(row=1, column=manuf_idx + i, value=col_name)

    # Fill values
    for r_idx, row in enumerate(merged_df.itertuples(index=False), start=2):
        for c_idx, col_name in enumerate(transfer_cols, start=mpn_idx + 1):
            ws.cell(row=r_idx, column=c_idx, value=getattr(row, col_name))

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    st.download_button(
        label="📥 Download Final Mapped File",
        data=buffer,
        file_name="ERPU2_MAPPED.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
