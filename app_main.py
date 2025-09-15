import streamlit as st
import pandas as pd
import math
from io import BytesIO
from openpyxl import load_workbook

# ✅ Keep template in project folder (no upload needed)
TEMPLATE_PATH = "Columbia Upload Excel Template_UPC_06182025 (1).xlsx"

st.title("📦 UPC & Polybag Sticker Order Generator")

# --- Upload Master Data (mandatory) ---
uploaded_file = st.file_uploader("Upload Master Data Excel", type=["xlsx"])
if uploaded_file:
    master_df = pd.read_excel(uploaded_file)
    master_df.columns = master_df.columns.str.strip()
    st.success("✅ Master Data uploaded successfully!")

    # --- Upload Gender Master Data (optional) ---
    gender_file = st.file_uploader("Upload Gender Master Data (Optional)", type=["xlsx"])
    gender_mapping = {}
    if gender_file:
        gender_df = pd.read_excel(gender_file)
        gender_df.columns = gender_df.columns.str.strip()

        style_col = None
        if "JDE Style" in gender_df.columns:
            style_col = "JDE Style"
        elif "Style" in gender_df.columns:
            style_col = "Style"

        if style_col and "Gender" in gender_df.columns:
            gender_mapping = dict(zip(gender_df[style_col].astype(str), gender_df["Gender"]))
            st.success("✅ Gender Master Data uploaded successfully!")
        else:
            st.error("❌ Gender Master must have columns: 'Style/JDE Style' and 'Gender'")

    # --- Buy Date Dropdown ---
    if "Buy Date" in master_df.columns:
        buy_date_options = master_df["Buy Date"].dropna().unique().tolist()
        selected_buy_date = st.selectbox("Select Buy Date", options=buy_date_options)
        master_df = master_df[master_df["Buy Date"] == selected_buy_date]
    else:
        st.error("❌ 'Buy Date' column not found in Master Data")
        st.stop()

    # --- Style Dropdown ---
    if "JDE Style" in master_df.columns:
        style_options = (
            master_df["JDE Style"]
            .dropna()
            .astype(str)
            .apply(lambda x: x[-4:])
            .unique()
            .tolist()
        )
        style_options.sort()
        selected_styles = st.multiselect("Select Style Number(s) (last 4 digits)", options=style_options)
    else:
        st.error("❌ 'JDE Style' column not found in Master Data")
        st.stop()

    # --- PDM Dropdown (for USA/INT only) ---
    pdm_options = ["071279", "073430", "096031", "121612", "122237", "123130"]
    selected_pdm = st.selectbox("Select UPC PDM (USA/INT)", options=pdm_options)

    # --- Sticker Qty Logic ---
    def calculate_sticker_qty(qty):
        try:
            qty = int(qty)
        except:
            qty = 0
        if qty < 50:
            return qty + 2
        else:
            return math.ceil(qty * 1.02)

    # --- Size order ---
    size_order = (
        [str(i) for i in range(2, 55)] +
        ["XXS","XS","S","M","L","XL","XXL",
         "1X","2X","3X","4X","5X","6X",
         "LT","XLT","2XT","3XT","4XT","5XT","6XT"]
    )

    def sort_by_size(df):
        if "Size" in df.columns:
            df["Size"] = pd.Categorical(df["Size"], categories=size_order, ordered=True)
        if "Inseam" in df.columns:
            df["Inseam"] = pd.Categorical(df["Inseam"], categories=size_order, ordered=True)
        return df.sort_values(
            by=["Item Number", "STYLE NUMBER", "COLOR NO", "Size", "Inseam"]
        ).reset_index(drop=True)

    # --- Prepare Output Rows ---
    output_rows_normal = []
    output_rows_japan = []

    for _, row in master_df.iterrows():
        style_full = str(row.get("JDE Style", ""))
        style_last4 = style_full[-4:] if len(style_full) >= 4 else style_full
        if style_last4 not in selected_styles:
            continue

        po = row.get("PO #", "")
        color = str(row.get("Color", "")).zfill(3)   # ✅ Force 3-digit color code
        size = str(row.get("F_Size", ""))            # ✅ Size now from F_Size
        f_dm = str(row.get("f_DM", ""))              # ✅ Inseam always from f_DM
        vendor = row.get("AB Number", "")
        season = row.get("Season", "")
        qty = row.get("Quantity", 0)
        coo = row.get("Country of Origin", "")
        country_full = str(row.get("Country", ""))

        # ✅ COO normalization
        if "india" in coo.lower():
            coo = "India"
        elif "bangladesh" in coo.lower():
            coo = "Bangladesh"

        # Destination mapping
        if country_full.upper() == "UNITED STATES":
            destination = "USA"
        elif country_full.upper() == "JAPAN":
            destination = "JAP"
        else:
            destination = "INT"

        sticker_qty = calculate_sticker_qty(qty)
        price = "YES" if destination == "USA" else "NO"

        gender_choice = gender_mapping.get(style_full, "Unisex")

        # --- Normal (USA/INT) ---
        if destination in ["USA", "INT"]:
            output_rows_normal.append({
                "Item Number": selected_pdm,
                "PO Number": po,
                "Quantity": sticker_qty,
                "Vendor": vendor,
                "VDATA": "",
                "DESTINATION": destination,
                "SEASON CODE": season,
                "STYLE NUMBER": style_full,
                "COLOR NO": color,
                "Barcode": "",
                "Country of Origin": coo,
                "GENDER": gender_choice,
                "Size": size,
                "Inseam": f_dm,
                "Price": price
            })
            output_rows_normal.append({
                "Item Number": "980010",
                "PO Number": po,
                "Quantity": sticker_qty,
                "Vendor": vendor,
                "VDATA": "",
                "DESTINATION": destination,
                "SEASON CODE": season,
                "STYLE NUMBER": style_full,
                "COLOR NO": color,
                "Barcode": "",
                "Country of Origin": coo,
                "GENDER": gender_choice,
                "Size": size,
                "Inseam": f_dm,
                "Price": "NO"
            })

        # --- Japan (separate file) ---
        elif destination == "JAP":
            # ✅ PDM 123138 → Price = YES
            output_rows_japan.append({
                "Item Number": "123138",   # Fixed UPC PDM for Japan
                "PO Number": po,
                "Quantity": sticker_qty,
                "Vendor": vendor,
                "VDATA": "",
                "DESTINATION": destination,
                "SEASON CODE": season,
                "STYLE NUMBER": style_full,
                "COLOR NO": color,
                "Barcode": "",
                "Country of Origin": coo,
                "GENDER": gender_choice,
                "Size": size,
                "Inseam": f_dm,
                "Price": "YES"
            })
            # ✅ Polybag stays NO
            output_rows_japan.append({
                "Item Number": "980010",   # Polybag fixed
                "PO Number": po,
                "Quantity": sticker_qty,
                "Vendor": vendor,
                "VDATA": "",
                "DESTINATION": destination,
                "SEASON CODE": season,
                "STYLE NUMBER": style_full,
                "COLOR NO": color,
                "Barcode": "",
                "Country of Origin": coo,
                "GENDER": gender_choice,
                "Size": size,
                "Inseam": f_dm,
                "Price": "NO"
            })

    # --- Convert to DataFrames ---
    output_df_normal = pd.DataFrame(output_rows_normal)
    output_df_normal = sort_by_size(output_df_normal)
    output_df_normal.insert(0, "S.No", range(1, len(output_df_normal) + 1))

    output_df_japan = pd.DataFrame(output_rows_japan)
    if not output_df_japan.empty:
        output_df_japan = sort_by_size(output_df_japan)
        output_df_japan.insert(0, "S.No", range(1, len(output_df_japan) + 1))

    # --- Preview ---
    st.subheader("Generated Order Form (Preview - USA/INT)")
    st.dataframe(output_df_normal)

    if not output_df_japan.empty:
        st.subheader("Generated Order Form (Preview - JAPAN)")
        st.dataframe(output_df_japan)

    # --- Function to write into template ---
    def generate_excel(df, filename):
        wb = load_workbook(TEMPLATE_PATH)
        ws = wb.active

        start_row = 7
        start_col = 2  # Column B

        for r_idx, row in df.iterrows():
            for c_idx, value in enumerate(row, start=start_col):
                ws.cell(row=start_row + r_idx, column=c_idx, value=row.iloc[c_idx - start_col])

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    # --- Download buttons ---
    if not output_df_normal.empty:
        filled_file_normal = generate_excel(output_df_normal, "order_form_normal.xlsx")
        st.download_button(
            label="📥 Download Order Form (USA/INT)",
            data=filled_file_normal,
            file_name="filled_order_form_USA_INT.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    if not output_df_japan.empty:
        filled_file_japan = generate_excel(output_df_japan, "order_form_japan.xlsx")
        st.download_button(
            label="📥 Download Order Form (JAPAN)",
            data=filled_file_japan,
            file_name="filled_order_form_JAPAN.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
